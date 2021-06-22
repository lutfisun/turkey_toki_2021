#!/usr/bin/python
# -*- coding: utf-8 -*-
file_name = 'lutfi.xlsx'
sheet = 'Sheet1'

import openpyxl
import urllib.request
from bs4 import BeautifulSoup
import pandas as pd
from lxml import etree
import xlsxwriter

months = {"Ocak": "1", "Şubat": "2", "Mart": "3", "Nisan": "4", "Mayıs": "5", "Haziran": "6", "Temmuz":"7", "Ağustos" : "8", "Eylül" : "9", "Ekim" : "10", "Kasım" : "11", "Aralık" : "12"}


def get_date_from_url(url):
    data = urllib.request.urlopen(url)  # it's a file like object and works just like a file
    found = False
    date = ''
    for line in data:  # files are iterable
        if date == '' and found and 'h4' in line.decode('utf8'
                ).strip('\n'):
            date = line.decode('utf8').strip('\n')
        if 'path' in line.decode('utf8').strip('\n'):
            found = True
    return date.strip().replace("<h4>","").replace("</h4>","")


def add_date_column(curr_row, new_date):
    if(isDate(new_date)):
        return curr_row.replace('</tr>', '<td>' + processDate(new_date)[2] + '</td><td>'+ processDate(new_date)[1] +'</td></tr>')
    else:
        return curr_row.replace('</tr>', '<td>' + " " + '</td><td>'+ " " +'</td></tr>')


def add_province (curr_row, province):
    return curr_row.replace('</tr>', '<td>' + province + '</td></tr>')

def add_date_title(curr_row):
    return curr_row.replace('</tr>', '<td>year</td><td>month</td><td>province</td></tr>')


# wb = openpyxl.load_workbook(file_name)
# ws = wb['Sheet1']
##print(ws.cell(row=2, column=1).hyperlink.target)

# fp = urllib.request.urlopen(ws.cell(row=28, column=2).hyperlink.target)

def isDate(date_value):
    if (any(char.isdigit() for char in date_value)) and date_value.split(" ")[1] in months:
        return True
    return False

def processDate(date_value):
    res = []
    day = date_value.split(" ")[0]
    month = months[date_value.split(" ")[1]]
    year = date_value.split(" ")[2]
    res.append(day)
    res.append(month)
    res.append(year)
    return res

def process_header(row):
    res_row = ""
    res = row.replace("<tr>","").replace("</tr>","")
    row_processor = BeautifulSoup(res)

    for t_data in row_processor.find_all('td'):
        res_row += str(t_data).replace("<td>","").replace("<br/>","").replace("</td>","").replace("\n","") + "[SECRET_SPLIT]"
    output_array = res_row.split("[SECRET_SPLIT]")
    return output_array[:len(output_array) - 1]

def process_row(row, url):
    res_row = ""
    res = row.replace("<tr>","").replace("</tr>","")
    row_processor = BeautifulSoup(res)

    for t_data in row_processor.find_all('td'):
        temp_row = str(t_data).replace("<td>","").replace("</td>","").replace("\n","") + "[SECRET_SPLIT]"
        temp_row = temp_row.replace("</a>","").replace("<a href=\""+url+"\">","").replace("<td style=\"text-align: left; padding-left:10px;\">","")
        res_row += temp_row
    output_array = res_row.split("[SECRET_SPLIT]")
    return output_array[:len(output_array) - 1]

def get_excel_file(array):
    #workbook = xlsxwriter.Workbook('arrays.xlsx')
    df = pd.DataFrame(array)
    df.to_excel(excel_writer = "test.xlsx", header=False, index=False)

def mainfunction():
    page = \
        urllib.request.urlopen('https://www.toki.gov.tr/proje-tipine-gore-uygulamalar'
                               ).read()
    soup = BeautifulSoup(page)  # Parse the HTML as a string

    # table = soup.find_all('table')[0] # Grab the first table

    new_tables = []
    elem = 0.0
    total = len(soup.find_all('table'))

    for table in soup.find_all('table'):
        print ('==================================================================================')
        first = True
        print(str(elem/total * 100) + " % complete.")
        for row in table.find_all('tr'):
            #print(str(row))
            new_row = ""
            new_row_array = []
            if first:
                new_row = add_date_title(str(row))
                first = False
                new_row_array = process_header(new_row)
                #print(process_header(new_row))
            else:
                date_link = "https://www.toki.gov.tr" + row.find_all('a')[0].get('href')
                date_value = get_date_from_url(date_link)
                new_row = add_date_column(str(row), date_value)
                new_row_array = process_row(new_row, row.find_all('a')[0].get('href'))
                new_row_array.append(new_row_array[1].split(" ")[0])

                #print(process_row(new_row, row.find_all('a')[0].get('href')))

            new_tables.append(new_row_array)
            print(len(new_row_array))
        new_tables.append([" new cat "] * 9)
        elem = elem + 1.0

    get_excel_file(new_tables)


mainfunction()
